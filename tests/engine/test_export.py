from pathlib import Path

from openpyxl import load_workbook

from dataclasses import replace

from engine.export import export_xlsx, export_pdf
from engine.io_json import (
    save_solution, load_solution, solution_to_dict, solution_from_dict,
    problem_to_dict, problem_from_dict, save_problem, load_problem,
)
from engine.solvers.greedy import GreedySolver


def test_export_xlsx_produces_readable_workbook(small_problem, tmp_path):
    sol = GreedySolver().solve(small_problem)
    out = tmp_path / "tt.xlsx"
    export_xlsx(sol, small_problem, out)
    assert out.exists() and out.stat().st_size > 0
    wb = load_workbook(out)
    # one sheet per division
    assert set(wb.sheetnames) == {d.id for d in small_problem.divisions}


def test_export_pdf_produces_file(small_problem, tmp_path):
    sol = GreedySolver().solve(small_problem)
    out = tmp_path / "tt.pdf"
    export_pdf(sol, small_problem, out)
    assert out.exists() and out.stat().st_size > 0


def test_solution_json_roundtrip(small_problem, tmp_path):
    sol = GreedySolver().solve(small_problem)
    path = tmp_path / "sol.json"
    save_solution(sol, path)
    loaded = load_solution(path)
    assert loaded.solver_name == sol.solver_name
    assert len(loaded.assignments) == len(sol.assignments)
    assert solution_to_dict(loaded) == solution_to_dict(sol)


def test_solution_dict_roundtrip(small_problem):
    sol = GreedySolver().solve(small_problem)
    assert solution_from_dict(solution_to_dict(sol)).assignments == sol.assignments


def test_problem_dict_roundtrip_is_stable(small_problem):
    # problem_to_dict must be an exact inverse of problem_from_dict: re-serializing the reparsed
    # instance yields byte-identical JSON, which is what makes a stored problem_snapshot reproducible.
    d1 = problem_to_dict(small_problem)
    d2 = problem_to_dict(problem_from_dict(d1))
    assert d1 == d2


def test_problem_dict_roundtrip_preserves_disruption_fields(small_problem):
    disrupted = replace(
        small_problem,
        blocked_slot_ids=frozenset({0, 1, 2}),
        relaxed_days=frozenset({1}),
        pinned_slots={"SESSION_X": 7},
    )
    reparsed = problem_from_dict(problem_to_dict(disrupted))
    assert reparsed.blocked_slot_ids == frozenset({0, 1, 2})
    assert reparsed.relaxed_days == frozenset({1})
    assert reparsed.pinned_slots == {"SESSION_X": 7}


def test_problem_json_file_roundtrip(small_problem, tmp_path):
    path = tmp_path / "problem.json"
    save_problem(small_problem, path)
    loaded = load_problem(path)
    assert problem_to_dict(loaded) == problem_to_dict(small_problem)

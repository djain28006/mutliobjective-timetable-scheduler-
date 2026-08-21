"""Export a Solution to human-shareable formats: Excel (per-division grid) and PDF."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from fpdf import FPDF
from fpdf.enums import XPos, YPos

from engine.models import ProblemInstance, Solution, expand_requirements

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def _division_grids(solution: Solution, problem: ProblemInstance) -> dict[str, dict[tuple[int, int], str]]:
    requirements = {r.id: r for r in expand_requirements(problem)}
    assignments = solution.assignment_by_session()
    slots_by_id = {t.id: t for t in problem.time_slots}
    rooms_by_id = problem.room_by_id()

    grids: dict[str, dict[tuple[int, int], str]] = {d.id: {} for d in problem.divisions}
    for req_id, assignment in assignments.items():
        req = requirements.get(req_id)
        if req is None or req.is_break:
            continue
        ts = slots_by_id.get(assignment.time_slot_id)
        if ts is None:
            continue
        room = rooms_by_id.get(assignment.room_id)
        label = f"{req.course_code} [{req.session_type.value}]"
        if req.faculty_id:
            label += f" {{{req.faculty_id}}}"
        if room:
            label += f" ({room.id})"
        if req.batch_id:
            label += f" {req.batch_id}"
        grid = grids.setdefault(req.division_id, {})
        for k in range(req.duration_slots):
            grid[(ts.day, ts.period + k)] = label

    # mark break slots
    for req_id, assignment in assignments.items():
        req = requirements.get(req_id)
        if req is None or not req.is_break:
            continue
        ts = slots_by_id.get(assignment.time_slot_id)
        if ts is None:
            continue
        grids.setdefault(req.division_id, {})[(ts.day, ts.period)] = "BREAK"

    return grids


def export_xlsx(solution: Solution, problem: ProblemInstance, path: str | Path) -> None:
    grids = _division_grids(solution, problem)
    periods = sorted({t.period for t in problem.time_slots})
    period_labels = {t.period: f"{t.start}-{t.end}" for t in problem.time_slots}

    wb = Workbook()
    wb.remove(wb.active)
    for division in problem.divisions:
        ws = wb.create_sheet(title=division.id)
        ws.cell(row=1, column=1, value="Period").font = Font(bold=True)
        for day in range(problem.days_per_week):
            cell = ws.cell(row=1, column=day + 2, value=DAY_NAMES[day])
            cell.font = Font(bold=True)

        grid = grids.get(division.id, {})
        for row_idx, period in enumerate(periods, start=2):
            ws.cell(row=row_idx, column=1, value=period_labels.get(period, str(period)))
            for day in range(problem.days_per_week):
                value = grid.get((day, period), "")
                cell = ws.cell(row=row_idx, column=day + 2, value=value)
                if value == "BREAK":
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for col_idx in range(1, problem.days_per_week + 2):
            ws.column_dimensions[chr(64 + col_idx)].width = 24

    wb.save(str(path))


def export_pdf(solution: Solution, problem: ProblemInstance, path: str | Path) -> None:
    grids = _division_grids(solution, problem)
    periods = sorted({t.period for t in problem.time_slots})
    period_labels = {t.period: f"{t.start}-{t.end}" for t in problem.time_slots}

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)

    for division in problem.divisions:
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, f"Timetable - Division {division.id}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        grid = grids.get(division.id, {})
        col_width = 260 / (problem.days_per_week + 1)
        row_height = 12

        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(col_width, row_height, "Period", border=1)
        for day in range(problem.days_per_week):
            pdf.cell(col_width, row_height, DAY_NAMES[day], border=1)
        pdf.ln(row_height)

        pdf.set_font("Helvetica", "", 7)
        for period in periods:
            pdf.cell(col_width, row_height, period_labels.get(period, str(period)), border=1)
            for day in range(problem.days_per_week):
                text = grid.get((day, period), "")
                pdf.cell(col_width, row_height, text[:28], border=1)
            pdf.ln(row_height)

    pdf.output(str(path))
